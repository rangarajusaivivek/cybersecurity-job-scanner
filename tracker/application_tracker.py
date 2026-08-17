"""
tracker/application_tracker.py
================================
Tracks all jobs found, scored, and applied to.
Saves to CSV (always) and Excel (optional).
Provides deduplication across daily runs.
Generates follow-up reminders.
"""

import os
import csv
import json
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
from utils.logger import log

TRACKER_PATH = "data/jobs_tracker.csv"
APPLIED_PATH = "data/applied_jobs.json"

COLUMNS = [
    "date_found", "id", "title", "company", "location", "salary",
    "source", "url", "apply_url", "score", "verdict",
    "opt_flag", "h1b_flag", "reasons", "talking_point",
    "status", "applied_date", "hm_name", "hm_email", "hm_linkedin",
    "follow_up_date", "notes", "interview_date", "offer_received",
]

STATUS_OPTIONS = ["New", "Applied", "Cold Email Sent", "Awaiting Response",
                  "Interview Scheduled", "Interviewed", "Offer Received",
                  "Rejected", "Withdrawn", "Following Up"]


def load_tracker() -> pd.DataFrame:
    """Load existing tracker or create empty one."""
    if os.path.exists(TRACKER_PATH):
        try:
            return pd.read_csv(TRACKER_PATH, dtype=str).fillna("")
        except Exception:
            pass
    df = pd.DataFrame(columns=COLUMNS)
    return df


def save_tracker(df: pd.DataFrame):
    """Save tracker to CSV."""
    os.makedirs("data", exist_ok=True)
    df.to_csv(TRACKER_PATH, index=False)
    log.info(f"Tracker saved: {TRACKER_PATH} ({len(df)} total records)")


def load_applied() -> set:
    """Load set of job IDs already applied to."""
    if os.path.exists(APPLIED_PATH):
        try:
            with open(APPLIED_PATH) as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()


def save_applied(applied_ids: set):
    """Save applied job IDs."""
    with open(APPLIED_PATH, "w", encoding="utf-8") as f:
        json.dump(list(applied_ids), f)


def add_jobs_to_tracker(new_jobs: list[dict]) -> tuple[int, int]:
    """
    Add new jobs to tracker. Skip duplicates.
    Returns (new_count, duplicate_count).
    """
    df = load_tracker()
    existing_ids = set(df["id"].tolist()) if "id" in df.columns else set()

    new_rows = []
    duplicates = 0
    today = datetime.now().strftime("%Y-%m-%d")

    for job in new_jobs:
        if job.get("id") in existing_ids:
            duplicates += 1
            continue

        row = {
            "date_found":    today,
            "id":            job.get("id", ""),
            "title":         job.get("title", ""),
            "company":       job.get("company", ""),
            "location":      job.get("location", ""),
            "salary":        job.get("salary", ""),
            "source":        job.get("source", ""),
            "url":           job.get("url", ""),
            "apply_url":     job.get("apply_url", job.get("url", "")),
            "score":         str(job.get("score", 0)),
            "verdict":       job.get("verdict", "INVESTIGATE"),
            "opt_flag":      str(job.get("opt_flag", False)),
            "h1b_flag":      str(job.get("h1b_flag", False)),
            "reasons":       " | ".join(job.get("reasons", [])) if isinstance(job.get("reasons"), list) else str(job.get("reasons", "")),
            "talking_point": job.get("talking_point", ""),
            "status":        "New",
            "applied_date":  "",
            "hm_name":       job.get("hm_name", ""),
            "hm_email":      job.get("hm_email", ""),
            "hm_linkedin":   job.get("hm_linkedin", ""),
            "follow_up_date": (datetime.now() + timedelta(days=5)).strftime("%Y-%m-%d"),
            "notes":         "",
            "interview_date": "",
            "offer_received": "",
        }
        new_rows.append(row)
        existing_ids.add(job.get("id"))

    if new_rows:
        new_df = pd.DataFrame(new_rows)
        df = pd.concat([df, new_df], ignore_index=True)
        save_tracker(df)

    log.info(f"Tracker: {len(new_rows)} new jobs added, {duplicates} duplicates skipped")
    return len(new_rows), duplicates


def get_follow_ups_due() -> list[dict]:
    """Return jobs where follow_up_date <= today and status is 'Applied' or 'Cold Email Sent'."""
    df = load_tracker()
    if df.empty:
        return []
    today = datetime.now().strftime("%Y-%m-%d")
    due = df[
        (df["follow_up_date"] <= today) &
        (df["status"].isin(["Applied", "Cold Email Sent", "Awaiting Response"]))
    ]
    return due.to_dict("records")


def get_stats() -> dict:
    """Return application statistics."""
    df = load_tracker()
    if df.empty:
        return {"total": 0}

    stats = {
        "total_found":     len(df),
        "apply_worthy":    len(df[df["verdict"] == "APPLY"]),
        "total_applied":   len(df[df["status"] == "Applied"]),
        "cold_emails":     len(df[df["status"] == "Cold Email Sent"]),
        "interviews":      len(df[df["status"].str.contains("Interview", na=False)]),
        "offers":          len(df[df["offer_received"] != ""]),
        "follow_ups_due":  len(get_follow_ups_due()),
        "top_sources":     df["source"].value_counts().head(5).to_dict(),
        "top_companies":   df[df["verdict"] == "APPLY"]["company"].value_counts().head(10).to_dict(),
        "avg_score":       round(pd.to_numeric(df["score"], errors="coerce").mean(), 1),
    }
    return stats


def export_to_excel(output_path: str = "data/jobs_tracker.xlsx"):
    """Export tracker to Excel with formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        df = load_tracker()
        if df.empty:
            return

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="All Jobs")

            wb = writer.book
            ws = wb["All Jobs"]

            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0A1628")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Color code by verdict
            verdict_col = list(df.columns).index("verdict") + 1
            for row in ws.iter_rows(min_row=2):
                verdict = row[verdict_col - 1].value
                if verdict == "APPLY":
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="E8F5E9")
                elif verdict == "SKIP":
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="F5F5F5")

            ws.freeze_panes = "A2"
            log.info(f"Excel exported: {output_path}")
    except Exception as e:
        log.warning(f"Excel export failed: {e}")
